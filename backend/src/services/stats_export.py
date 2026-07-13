# -*- coding: utf-8 -*-
"""Экспорт корпусной статистики в CSV и Excel."""
from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass
from datetime import datetime
from typing import Iterable, Sequence

from openpyxl import Workbook

from backend.src.services.audio_filter import StatsFilters
from backend.src.services.word_stats import (
    DateWordsResult,
    FrequentWordsResult,
    LanguageWordsResult,
    SpeakerWordsResult,
)

_LANG_LABELS = {
    "ru": "Русский",
    "tt": "Татарский",
    "unknown": "Другие",
}


@dataclass
class StatsExportSection:
    category_title: str
    sheet_title: str
    filters: StatsFilters
    headers: Sequence[str]
    data_rows: list[Sequence[object]]
    audio_labels: Sequence[str] | None = None
    limit: int | None = None
    summary_rows: Sequence[Sequence[str]] | None = None


def _sanitize_sheet_title(title: str) -> str:
    cleaned = re.sub(r"[\\/*?:\[\]]", " ", title).strip()
    return (cleaned or "Статистика")[:31]


def _format_percent(count: int, total: int) -> str:
    if total <= 0:
        return "0"
    return f"{(count / total) * 100:.2f}"


def _build_filter_rows(
    category_title: str,
    filters: StatsFilters,
    *,
    audio_labels: Sequence[str] | None = None,
    limit: int | None = None,
    summary_rows: Sequence[Sequence[str]] | None = None,
) -> list[list[str]]:
    rows: list[list[str]] = [["Категория", category_title]]

    if filters.langs:
        labels = [_LANG_LABELS.get(lang, lang) for lang in filters.langs]
        rows.append(["Язык", ", ".join(labels)])
    if filters.speakers:
        rows.append(["Говорящие", ", ".join(filters.speakers)])
    if filters.date_from:
        rows.append(["Дата с", filters.date_from])
    if filters.date_to:
        rows.append(["Дата по", filters.date_to])
    if filters.audio_ids:
        if audio_labels:
            rows.append(["Аудиозаписи", "; ".join(audio_labels)])
        else:
            rows.append(["Аудиозаписи", str(len(filters.audio_ids))])
    if limit is not None:
        rows.append(["Лимит строк", str(limit)])
    if summary_rows:
        rows.extend(summary_rows)

    rows.append([])
    return rows


def _write_table(
    writer: csv.writer,
    headers: Sequence[str],
    data_rows: Iterable[Sequence[object]],
) -> None:
    writer.writerow(headers)
    for row in data_rows:
        writer.writerow(row)


def _write_xlsx_table(
    worksheet,
    start_row: int,
    headers: Sequence[str],
    data_rows: Iterable[Sequence[object]],
) -> int:
    row_idx = start_row
    for col_idx, header in enumerate(headers, start=1):
        worksheet.cell(row=row_idx, column=col_idx, value=header)
    row_idx += 1
    for row in data_rows:
        for col_idx, value in enumerate(row, start=1):
            worksheet.cell(row=row_idx, column=col_idx, value=value)
        row_idx += 1
    return row_idx


def build_export_payload(
    *,
    export_format: str,
    category_slug: str,
    category_title: str,
    filters: StatsFilters,
    headers: Sequence[str],
    data_rows: list[Sequence[object]],
    audio_labels: Sequence[str] | None = None,
    limit: int | None = None,
    summary_rows: Sequence[Sequence[str]] | None = None,
) -> tuple[bytes, str, str]:
    normalized_format = export_format.strip().lower()
    if normalized_format not in {"csv", "xlsx"}:
        raise ValueError("Поддерживаются только форматы csv и xlsx")

    filter_rows = _build_filter_rows(
        category_title,
        filters,
        audio_labels=audio_labels,
        limit=limit,
        summary_rows=summary_rows,
    )

    if normalized_format == "csv":
        buffer = io.StringIO()
        writer = csv.writer(buffer, lineterminator="\n")
        for row in filter_rows:
            writer.writerow(row)
        _write_table(writer, headers, data_rows)
        content = buffer.getvalue().encode("utf-8-sig")
        media_type = "text/csv; charset=utf-8"
        extension = "csv"
    else:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Статистика"
        row_idx = 1
        for row in filter_rows:
            if not row:
                row_idx += 1
                continue
            worksheet.cell(row=row_idx, column=1, value=row[0])
            if len(row) > 1:
                worksheet.cell(row=row_idx, column=2, value=row[1])
            row_idx += 1
        _write_xlsx_table(worksheet, row_idx, headers, data_rows)
        output = io.BytesIO()
        workbook.save(output)
        content = output.getvalue()
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        extension = "xlsx"

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"stats_{category_slug}_{timestamp}.{extension}"
    return content, media_type, filename


def export_frequent_words(
    result: FrequentWordsResult,
    filters: StatsFilters,
    *,
    export_format: str,
    audio_labels: Sequence[str] | None = None,
) -> tuple[bytes, str, str]:
    total = result.total_words
    data_rows = [
        (
            item.text,
            _LANG_LABELS.get(item.language, item.language),
            item.count,
            _format_percent(item.count, total),
        )
        for item in result.items
    ]
    return build_export_payload(
        export_format=export_format,
        category_slug="frequent_words",
        category_title="Самые частые слова",
        filters=filters,
        headers=["Слово", "Язык", "Количество", "Доля (%)"],
        data_rows=data_rows,
        audio_labels=audio_labels,
        limit=result.limit,
        summary_rows=[
            ["Всего слов", str(result.total_words)],
            ["Уникальных слов", str(result.unique_words)],
        ],
    )


def export_language_stats(
    result: LanguageWordsResult,
    filters: StatsFilters,
    *,
    export_format: str,
    audio_labels: Sequence[str] | None = None,
) -> tuple[bytes, str, str]:
    total = result.total_words
    data_rows = [
        (item.label, item.count, _format_percent(item.count, total))
        for item in result.items
    ]
    return build_export_payload(
        export_format=export_format,
        category_slug="languages",
        category_title="Статистика по языкам",
        filters=filters,
        headers=["Язык", "Количество", "Доля (%)"],
        data_rows=data_rows,
        audio_labels=audio_labels,
        summary_rows=[["Всего слов", str(result.total_words)]],
    )


def export_date_stats(
    result: DateWordsResult,
    filters: StatsFilters,
    *,
    export_format: str,
    audio_labels: Sequence[str] | None = None,
) -> tuple[bytes, str, str]:
    total = result.total_words
    data_rows = [
        (item.label, item.count, _format_percent(item.count, total))
        for item in result.items
    ]
    return build_export_payload(
        export_format=export_format,
        category_slug="dates",
        category_title="Статистика по датам",
        filters=filters,
        headers=["Дата", "Количество", "Доля (%)"],
        data_rows=data_rows,
        audio_labels=audio_labels,
        limit=result.limit,
        summary_rows=[
            ["Всего слов", str(result.total_words)],
            ["Дат", str(result.total_dates)],
        ],
    )


def export_speaker_stats(
    result: SpeakerWordsResult,
    filters: StatsFilters,
    *,
    export_format: str,
    audio_labels: Sequence[str] | None = None,
) -> tuple[bytes, str, str]:
    total = result.total_words
    data_rows = [
        (item.label, item.count, _format_percent(item.count, total))
        for item in result.items
    ]
    return build_export_payload(
        export_format=export_format,
        category_slug="speakers",
        category_title="Статистика по говорящим",
        filters=filters,
        headers=["Говорящий", "Количество", "Доля (%)"],
        data_rows=data_rows,
        audio_labels=audio_labels,
        limit=result.limit,
        summary_rows=[
            ["Всего слов", str(result.total_words)],
            ["Говорящих", str(result.total_speakers)],
        ],
    )


def _build_frequent_words_section(
    result: FrequentWordsResult,
    filters: StatsFilters,
    *,
    audio_labels: Sequence[str] | None = None,
) -> StatsExportSection:
    total = result.total_words
    return StatsExportSection(
        category_title="Самые частые слова",
        sheet_title="Частые слова",
        filters=filters,
        headers=["Слово", "Язык", "Количество", "Доля (%)"],
        data_rows=[
            (
                item.text,
                _LANG_LABELS.get(item.language, item.language),
                item.count,
                _format_percent(item.count, total),
            )
            for item in result.items
        ],
        audio_labels=audio_labels,
        limit=result.limit,
        summary_rows=[
            ["Всего слов", str(result.total_words)],
            ["Уникальных слов", str(result.unique_words)],
        ],
    )


def _build_language_section(
    result: LanguageWordsResult,
    filters: StatsFilters,
    *,
    audio_labels: Sequence[str] | None = None,
) -> StatsExportSection:
    total = result.total_words
    return StatsExportSection(
        category_title="Статистика по языкам",
        sheet_title="Языки",
        filters=filters,
        headers=["Язык", "Количество", "Доля (%)"],
        data_rows=[
            (item.label, item.count, _format_percent(item.count, total))
            for item in result.items
        ],
        audio_labels=audio_labels,
        summary_rows=[["Всего слов", str(result.total_words)]],
    )


def _build_date_section(
    result: DateWordsResult,
    filters: StatsFilters,
    *,
    audio_labels: Sequence[str] | None = None,
) -> StatsExportSection:
    total = result.total_words
    return StatsExportSection(
        category_title="Статистика по датам",
        sheet_title="Даты",
        filters=filters,
        headers=["Дата", "Количество", "Доля (%)"],
        data_rows=[
            (item.label, item.count, _format_percent(item.count, total))
            for item in result.items
        ],
        audio_labels=audio_labels,
        limit=result.limit,
        summary_rows=[
            ["Всего слов", str(result.total_words)],
            ["Дат", str(result.total_dates)],
        ],
    )


def _build_speaker_section(
    result: SpeakerWordsResult,
    filters: StatsFilters,
    *,
    audio_labels: Sequence[str] | None = None,
) -> StatsExportSection:
    total = result.total_words
    return StatsExportSection(
        category_title="Статистика по говорящим",
        sheet_title="Говорящие",
        filters=filters,
        headers=["Говорящий", "Количество", "Доля (%)"],
        data_rows=[
            (item.label, item.count, _format_percent(item.count, total))
            for item in result.items
        ],
        audio_labels=audio_labels,
        limit=result.limit,
        summary_rows=[
            ["Всего слов", str(result.total_words)],
            ["Говорящих", str(result.total_speakers)],
        ],
    )


def _write_section_to_csv(writer: csv.writer, section: StatsExportSection) -> None:
    writer.writerow(["Раздел", section.category_title])
    filter_rows = _build_filter_rows(
        section.category_title,
        section.filters,
        audio_labels=section.audio_labels,
        limit=section.limit,
        summary_rows=section.summary_rows,
    )
    for row in filter_rows:
        writer.writerow(row)
    _write_table(writer, section.headers, section.data_rows)
    writer.writerow([])


def _write_section_to_xlsx(worksheet, section: StatsExportSection) -> None:
    row_idx = 1
    filter_rows = _build_filter_rows(
        section.category_title,
        section.filters,
        audio_labels=section.audio_labels,
        limit=section.limit,
        summary_rows=section.summary_rows,
    )
    for row in filter_rows:
        if not row:
            row_idx += 1
            continue
        worksheet.cell(row=row_idx, column=1, value=row[0])
        if len(row) > 1:
            worksheet.cell(row=row_idx, column=2, value=row[1])
        row_idx += 1
    _write_xlsx_table(worksheet, row_idx, section.headers, section.data_rows)


def export_all_stats(
    *,
    frequent_words: FrequentWordsResult,
    frequent_filters: StatsFilters,
    languages: LanguageWordsResult,
    language_filters: StatsFilters,
    dates: DateWordsResult,
    date_filters: StatsFilters,
    speakers: SpeakerWordsResult,
    speaker_filters: StatsFilters,
    export_format: str,
    frequent_audio_labels: Sequence[str] | None = None,
    language_audio_labels: Sequence[str] | None = None,
    date_audio_labels: Sequence[str] | None = None,
    speaker_audio_labels: Sequence[str] | None = None,
) -> tuple[bytes, str, str]:
    normalized_format = export_format.strip().lower()
    if normalized_format not in {"csv", "xlsx"}:
        raise ValueError("Поддерживаются только форматы csv и xlsx")

    sections = [
        _build_frequent_words_section(
            frequent_words,
            frequent_filters,
            audio_labels=frequent_audio_labels,
        ),
        _build_language_section(
            languages,
            language_filters,
            audio_labels=language_audio_labels,
        ),
        _build_date_section(
            dates,
            date_filters,
            audio_labels=date_audio_labels,
        ),
        _build_speaker_section(
            speakers,
            speaker_filters,
            audio_labels=speaker_audio_labels,
        ),
    ]

    if normalized_format == "csv":
        buffer = io.StringIO()
        writer = csv.writer(buffer, lineterminator="\n")
        for section in sections:
            _write_section_to_csv(writer, section)
        content = buffer.getvalue().encode("utf-8-sig")
        media_type = "text/csv; charset=utf-8"
        extension = "csv"
    else:
        workbook = Workbook()
        first_sheet = True
        for section in sections:
            if first_sheet:
                worksheet = workbook.active
                worksheet.title = _sanitize_sheet_title(section.sheet_title)
                first_sheet = False
            else:
                worksheet = workbook.create_sheet(title=_sanitize_sheet_title(section.sheet_title))
            _write_section_to_xlsx(worksheet, section)
        output = io.BytesIO()
        workbook.save(output)
        content = output.getvalue()
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        extension = "xlsx"

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"stats_all_{timestamp}.{extension}"
    return content, media_type, filename

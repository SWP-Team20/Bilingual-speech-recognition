# -*- coding: utf-8 -*-
"""Юнит-тесты экспорта корпусной статистики."""
import csv
import io

from uuid import uuid4

import pytest
from openpyxl import load_workbook

from backend.src.services.audio_filter import StatsFilters
from backend.src.services.stats_export import (
    build_export_payload,
    export_all_stats,
    export_date_stats,
    export_frequent_words,
    export_language_stats,
    export_speaker_stats,
)
from backend.src.services.word_stats import (
    DateWordCount,
    DateWordsResult,
    FrequentWordsResult,
    LanguageWordCount,
    LanguageWordsResult,
    SpeakerWordCount,
    SpeakerWordsResult,
    WordFrequency,
)

pytestmark = pytest.mark.unit


def _read_csv_rows(content: bytes) -> list[list[str]]:
    text = content.decode("utf-8-sig")
    return list(csv.reader(io.StringIO(text)))


def test_build_export_payload_csv_includes_filters_and_data():
    audio_id = uuid4()
    filters = StatsFilters(
        langs=["ru", "tt"],
        speakers=["мама"],
        date_from="2026-01-01",
        date_to="2026-01-31",
        audio_ids=[audio_id],
    )
    content, media_type, filename = build_export_payload(
        export_format="csv",
        category_slug="languages",
        category_title="Статистика по языкам",
        filters=filters,
        headers=["Язык", "Количество"],
        data_rows=[("Русский", 10)],
        audio_labels=["sample.wav"],
        summary_rows=[["Всего слов", "10"]],
    )

    assert media_type == "text/csv; charset=utf-8"
    assert filename.endswith(".csv")
    rows = _read_csv_rows(content)
    assert rows[0] == ["Категория", "Статистика по языкам"]
    assert ["Язык", "Русский, Татарский"] in rows
    assert ["Говорящие", "мама"] in rows
    assert ["Дата с", "2026-01-01"] in rows
    assert ["Дата по", "2026-01-31"] in rows
    assert ["Аудиозаписи", "sample.wav"] in rows
    assert ["Всего слов", "10"] in rows
    assert ["Язык", "Количество"] in rows
    assert ["Русский", "10"] in rows


def test_export_frequent_words_xlsx_contains_table():
    result = FrequentWordsResult(
        items=[WordFrequency(text="дом", language="ru", count=5)],
        total_words=5,
        unique_words=1,
        limit=30,
    )
    filters = StatsFilters(langs=["ru"])

    content, media_type, filename = export_frequent_words(
        result,
        filters,
        export_format="xlsx",
    )

    assert media_type.endswith("spreadsheetml.sheet")
    assert filename.endswith(".xlsx")

    workbook = load_workbook(io.BytesIO(content))
    worksheet = workbook.active
    assert worksheet["A1"].value == "Категория"
    assert worksheet["B1"].value == "Самые частые слова"
    assert worksheet["A2"].value == "Язык"
    assert worksheet["B2"].value == "Русский"

    values = [worksheet.cell(row=row, column=1).value for row in range(1, worksheet.max_row + 1)]
    assert "Слово" in values
    assert "дом" in values


def test_export_language_stats_csv_percent_column():
    result = LanguageWordsResult(
        items=[
            LanguageWordCount(language="ru", label="Русский", count=75),
            LanguageWordCount(language="tt", label="Татарский", count=25),
        ],
        total_words=100,
    )
    content, _, _ = export_language_stats(result, StatsFilters(), export_format="csv")
    rows = _read_csv_rows(content)
    assert ["Русский", "75", "75.00"] in rows
    assert ["Татарский", "25", "25.00"] in rows


def test_export_date_and_speaker_stats_reject_unknown_format():
    date_result = DateWordsResult(
        items=[DateWordCount(date="2026-01-01", label="2026-01-01", count=3)],
        total_words=3,
        total_dates=1,
        limit=30,
    )
    speaker_result = SpeakerWordsResult(
        items=[SpeakerWordCount(speaker_id=1, label="мама", count=7)],
        total_words=7,
        total_speakers=1,
        limit=20,
    )

    with pytest.raises(ValueError, match="csv и xlsx"):
        export_date_stats(date_result, StatsFilters(), export_format="pdf")

    with pytest.raises(ValueError, match="csv и xlsx"):
        export_speaker_stats(speaker_result, StatsFilters(), export_format="ods")


def test_export_all_stats_xlsx_has_four_sheets():
    frequent = FrequentWordsResult(
        items=[WordFrequency(text="дом", language="ru", count=2)],
        total_words=2,
        unique_words=1,
        limit=30,
    )
    languages = LanguageWordsResult(
        items=[LanguageWordCount(language="ru", label="Русский", count=2)],
        total_words=2,
    )
    dates = DateWordsResult(
        items=[DateWordCount(date="2026-01-01", label="2026-01-01", count=2)],
        total_words=2,
        total_dates=1,
        limit=30,
    )
    speakers = SpeakerWordsResult(
        items=[SpeakerWordCount(speaker_id=1, label="мама", count=2)],
        total_words=2,
        total_speakers=1,
        limit=20,
    )

    content, media_type, filename = export_all_stats(
        frequent_words=frequent,
        frequent_filters=StatsFilters(langs=["ru"]),
        languages=languages,
        language_filters=StatsFilters(),
        dates=dates,
        date_filters=StatsFilters(),
        speakers=speakers,
        speaker_filters=StatsFilters(),
        export_format="xlsx",
    )

    assert media_type.endswith("spreadsheetml.sheet")
    assert filename.endswith(".xlsx")

    workbook = load_workbook(io.BytesIO(content))
    assert workbook.sheetnames == ["Частые слова", "Языки", "Даты", "Говорящие"]


def test_export_all_stats_csv_contains_all_sections():
    frequent = FrequentWordsResult(
        items=[WordFrequency(text="дом", language="ru", count=2)],
        total_words=2,
        unique_words=1,
        limit=30,
    )
    languages = LanguageWordsResult(
        items=[LanguageWordCount(language="ru", label="Русский", count=2)],
        total_words=2,
    )
    dates = DateWordsResult(
        items=[DateWordCount(date="2026-01-01", label="2026-01-01", count=2)],
        total_words=2,
        total_dates=1,
        limit=30,
    )
    speakers = SpeakerWordsResult(
        items=[SpeakerWordCount(speaker_id=1, label="мама", count=2)],
        total_words=2,
        total_speakers=1,
        limit=20,
    )

    content, media_type, filename = export_all_stats(
        frequent_words=frequent,
        frequent_filters=StatsFilters(langs=["ru"]),
        languages=languages,
        language_filters=StatsFilters(speakers=["мама"]),
        dates=dates,
        date_filters=StatsFilters(),
        speakers=speakers,
        speaker_filters=StatsFilters(),
        export_format="csv",
    )

    assert media_type == "text/csv; charset=utf-8"
    assert filename.endswith(".csv")

    rows = _read_csv_rows(content)
    assert ["Раздел", "Самые частые слова"] in rows
    assert ["Раздел", "Статистика по языкам"] in rows
    assert ["Раздел", "Статистика по датам"] in rows
    assert ["Раздел", "Статистика по говорящим"] in rows
    assert ["Говорящие", "мама"] in rows
